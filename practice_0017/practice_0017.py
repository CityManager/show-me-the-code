#!/usr/bin/env python3
# _*_ coding:utf-8 _*_

__author__ = 'CityManager'

"""
http://openpyxl.readthedocs.org
"""

from openpyxl import load_workbook
from xml.dom.minidom import getDOMImplementation


class Student(object):
    def __init__(self, sid, info):
        self.sid = sid
        self.name = info[0]
        self.math_score = info[1]
        self.chinese_score = info[2]
        self.english_score = info[3]

    def __str__(self):
        return '{}: {}'.format(self.sid, [self.name, self.math_score, self.chinese_score, self.english_score])


def read_excel(filename, sheet_name=None, range_cell='B1:F3'):
    students = list()
    wb = load_workbook(filename)
    ws = wb[sheet_name] if sheet_name else wb.active
    rows = ws.iter_rows(range_cell)
    for row in rows:
        students.append(Student(row[0].value, [cell.value for cell in row[1:]]))
    students.sort(key=lambda student: student.sid)
    return students


def save_xml(filename, students_list):
    impl = getDOMImplementation()
    dom = impl.createDocument(None, None, None)

    root = dom.createElement('root')
    students = dom.createElement('students')
    root.appendChild(students)
    comment_content = '\n    学生信息表\n    "id" : [名字, 数学, 语文, 英文]\n'
    comment = dom.createComment(comment_content)
    text = '{\n' + ",\n".join(['    %s' % stu for stu in students_list]) + '\n}'
    text_node = dom.createTextNode(text)
    students.appendChild(comment)
    students.appendChild(text_node)

    dom.appendChild(root)

    with open(filename, 'w', encoding='utf-8') as f:
        dom.writexml(f, newl='\n', encoding='utf-8')


if __name__ == '__main__':
    students_data = read_excel('student.xlsx', 'student', 'B1:F3')
    save_xml('student.xml', students_data)
